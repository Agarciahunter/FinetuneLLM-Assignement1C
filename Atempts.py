#conda activate arc1c
from peft import PeftModel, PeftConfig
from datasets import load_dataset
import random
from codebleu import calc_codebleu
from rouge import Rouge
from bert_score import score
import evaluate
import openpyxl
from transformers import (
    AutoModelForCausalLM,
     AutoTokenizer)

def getOutput(tokenizer,model,testPrompt,hparam,size=0):
    input = tokenizer(testPrompt, return_tensors="pt").input_ids

    if hparam == "vanilla":
        # Generate output using vanilla decoding
        outputs = model.generate(input, max_length = 450)
    elif hparam == "topK":
        # Generate output using top-K sampling
        outputs = model.generate(input,
                                 max_length = 450,
                                 do_sample=True,
                                 top_k=size)
    elif hparam == "beam":
        # Generate output using beam search
        outputs = model.generate(input,
                                 max_length = 450,
                                 num_beams=size,
                                 early_stopping=True)
    elif hparam == "temp":
         # Generate output using temperature sampling
         outputs = model.generate(input,
                                 max_length = 450,
                                 do_sample=True,
                                 top_k=0,
                                 temperature = size)

    text = tokenizer.decode(outputs[0], skip_special_tokens=True)
    return text


modelList = [
    "./Llama",
    # "./FTPhi2",
    # "./Mistral"
]

outputType = [
    "vanilla",
    "topK",
    # "beam",
    # "temp"
]

topKsize = [
    2,
    4,
    6,
    8
]

beamsize = [
    2,
    3,
    4,
    5
]

tempSize = [
    .1,
    .25,
    .5,
    .75
]

datapath = "flytech/python-codes-25k"

# Load the dataset
dataset = load_dataset("flytech/python-codes-25k", split='train').train_test_split(test_size=0.001, train_size=0.01)
num_samples = 20
input_instructions = dataset["test"]["instruction"]
ground_truth = dataset["test"]["output"]
# Load dataset
dataset = load_dataset("flytech/python-codes-25k", split='train')
numInputs = 3

randrows = []
for i in range(numInputs):
    randrows.append(random.randint(0,len(dataset)))
# Choose random samples from the instructions
random_indices = random.sample(range(len(input_instructions)), num_samples)

# Select random rows from the dataset
dataset = dataset.select(randrows)

# Open the Excel file
workbook = openpyxl.load_workbook('Evaluation_Results.xlsx')

# Access the specific sheets
metrics_sheet = workbook['Metrics']
hyperparam_sheet = workbook['Hyperparam']

# Set starting row and column for Metric sheet
start_row1 = 2
start_column1 = 2  # Column B
# Set starting row and column for Metric sheet
start_row2 = 2
start_column2 = 3

for modelpath in modelList:
    # Load model and tokenizer
    model = AutoModelForCausalLM.from_pretrained(modelpath)
    model = PeftModel.from_pretrained(model, modelpath)
    tokenizer = AutoTokenizer.from_pretrained(modelpath)

    for hparam in outputType:
        sizes = []
        if hparam == "vanilla":
            sizes = [1]
        elif hparam == "topK":
            sizes = topKsize.copy()
        elif hparam == "beam":
            sizes = beamsize.copy()
        elif hparam == "temp":
            sizes = tempSize.copy()

        for size in sizes:
            referencelist = []
            predictionlist = []
            for i in range(numInputs):
                print("Getting output for: " + str(modelpath)+ "... output type: "+ str(hparam)+ " size = "+ str(size) + "...Instruction:" + str(i+1))
                testPrompt = dataset[i]["instruction"]
                
                text = getOutput(tokenizer,model,testPrompt,hparam,size)
                
                referencelist.append(dataset[i]["output"])
                predictionlist.append(text)
            
            print("Results for " + str(modelpath)+ "... output type: "+ str(hparam)+ " size = "+ str(size))
            print('-' * 80)
            
            # Calculate BLEU score alongside CodeBLEU
            bleu = evaluate.load("bleu")
            bleu_score = bleu.compute(predictions=predictionlist, references=referencelist)
            print("BLEU Score: " + str(bleu_score["bleu"]))

            ##codebleu##
            codebleuResult = calc_codebleu(referencelist, predictionlist, lang="python", weights=(0.25, 0.25, 0.25, 0.25), tokenizer=None)
            print("CodeBleu Scrore: " + str(codebleuResult["codebleu"]))
            ##rouge##
            rouge = Rouge()
            scores = rouge.get_scores(predictionlist, referencelist, avg=True)
            print("Rouge-L score: " + str(scores["rouge-l"]))
            ##BERTscore##
            P, R, F1 = score(predictionlist, referencelist, lang="en", verbose=True)
            print("BERTScore:")
            print(P, R, F1)
            bert_score_str = f"r: {R.item()}, p: {P.item()}, f: {F1.item()}"

            print('-' * 80)
            print("")

            # Initialize variables to store cumulative scores
            total_compilability_score = 0
            total_correctness_score = 0
            print("For Human Evaluation on : " + str(modelpath)+ "... output type: "+ str(hparam)+ " size = "+ str(size))
            for idx in random_indices:
                prompt = input_instructions[idx]
                expected_output = ground_truth[idx]

                input_ids = tokenizer(prompt, return_tensors="pt").input_ids
                outputs = model.generate(input_ids, max_length=200)
                generated_code = tokenizer.decode(outputs[0], skip_special_tokens=True)
                # Print instruction, generated text, and expected output for reference
                print("Instruction:", prompt)
                print()
                print("Generated Code:")
                print(generated_code)
                print()
                print("Expected Output:")
                print(expected_output)
                # Simulate human evaluation scores (replace with actual evaluation process)
                compilability_score = float(input(f"Rate the compilability (0-1) of the generated code for instruction: '{prompt}': "))
                correctness_score = float(input(f"Rate the functional correctness (0-1) of the generated code for instruction: '{prompt}': "))
                # Accumulate scores
                total_compilability_score += compilability_score
                total_correctness_score += correctness_score
                print("=" * 50)
            # Calculate average scores
            average_compilability_score = total_compilability_score / num_samples
            average_correctness_score = total_correctness_score / num_samples
            # Calculate final human evaluation score based on the average of compilability and correctness scores
            final_human_evaluation_score = (average_compilability_score + average_correctness_score) / 2
            print("Final Human Evaluation Score:", final_human_evaluation_score)

            # Save results to appropriate sheet
            if hparam == "vanilla":
                sheet = metrics_sheet
                # Append results to Metrics sheet starting from row 3, column C
                sheet.cell(row=start_row1, column=start_column1).value = bleu_score["bleu"]
                sheet.cell(row=start_row1, column=start_column1+1).value = str(scores["rouge-l"])
                sheet.cell(row=start_row1, column=start_column1+2).value = bert_score_str
                sheet.cell(row=start_row1, column=start_column1+3).value = codebleuResult["codebleu"]
                sheet.cell(row=start_row1, column=start_column1+4).value = final_human_evaluation_score
                # Increment start_row for the next iteration
                start_row1 += 1
            else:
                sheet = hyperparam_sheet
                # Append results to Hyperparam sheet starting from row 3, column C
                sheet.cell(row=start_row2, column=start_column2).value = size
                sheet.cell(row=start_row2, column=start_column2+1).value = bleu_score["bleu"]
                sheet.cell(row=start_row2, column=start_column2+2).value = str(scores["rouge-l"])
                sheet.cell(row=start_row2, column=start_column2+3).value = bert_score_str
                sheet.cell(row=start_row2, column=start_column2+4).value = codebleuResult["codebleu"]
                sheet.cell(row=start_row1, column=start_column2+5).value = final_human_evaluation_score
                # Increment start_row for the next iteration
                start_row2 += 1
            
# Save the changes to the Excel file
workbook.save('Evaluation_Results.xlsx')